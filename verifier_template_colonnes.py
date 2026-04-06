"""
Verification de la structure du template Excel
"""
from openpyxl import load_workbook

wb = load_workbook('py_backend/Liasse_officielle_revise.xlsx')

# Verifier l'onglet ACTIF
if 'ACTIF' in wb.sheetnames:
    ws = wb['ACTIF']
    print('=== ONGLET ACTIF ===')
    print('Structure des colonnes (ligne 10):')
    for col in ['D', 'E', 'F', 'G', 'H', 'I']:
        cell_value = ws[f'{col}10'].value
        print(f'  Colonne {col}: {cell_value}')
    
    print('\nLigne AI (Terrains) - recherche:')
    for row in range(10, 30):
        if ws[f'A{row}'].value == 'AI':
            print(f'  Ligne {row}: REF=AI')
            for col in ['D', 'E', 'F', 'G', 'H', 'I']:
                print(f'    {col}: {ws[f"{col}{row}"].value}')
            break

# Verifier l'onglet TFT
tft_onglet = None
for name in wb.sheetnames:
    if 'TFT' in name.upper():
        tft_onglet = name
        break

if tft_onglet:
    ws_tft = wb[tft_onglet]
    print(f'\n=== ONGLET {tft_onglet} ===')
    print('Structure des colonnes (ligne 10):')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        cell_value = ws_tft[f'{col}10'].value
        print(f'  Colonne {col}: {cell_value}')
    
    print('\nLigne ZA - recherche:')
    for row in range(10, 50):
        if ws_tft[f'A{row}'].value == 'ZA':
            print(f'  Ligne {row}: REF=ZA')
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                print(f'    {col}: {ws_tft[f"{col}{row}"].value}')
            break

wb.close()
