#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Test direct de l'export sans passer par FastAPI
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import sys

print("="*80)
print("TEST DIRECT - EXPORT LIASSE")
print("="*80)

try:
    # Charger les balances
    print("\n📂 Chargement des balances...")
    balance_n = pd.read_excel("P000 -BALANCE DEMO N_N-1_N-2.xls", sheet_name="BALANCE N")
    balance_n1 = pd.read_excel("P000 -BALANCE DEMO N_N-1_N-2.xls", sheet_name="BALANCE N-1")
    balance_n2 = pd.read_excel("P000 -BALANCE DEMO N_N-1_N-2.xls", sheet_name="BALANCE N-2")
    print(f"   ✅ Balances chargées: N={len(balance_n)}, N-1={len(balance_n1)}, N-2={len(balance_n2)}")
    
    # Charger les correspondances
    print("\n📂 Chargement des correspondances...")
    with open('correspondances_syscohada.json', 'r', encoding='utf-8') as f:
        correspondances = json.load(f)
    print("   ✅ Correspondances chargées")
    
    # Importer la fonction de génération
    print("\n📊 Import de la fonction de génération...")
    from etats_financiers_v2 import process_balance_to_liasse_format
    print("   ✅ Fonction importée")
    
    # Générer les états
    print("\n📊 Génération des états financiers...")
    results = process_balance_to_liasse_format(
        balance_n, balance_n1, balance_n2, correspondances
    )
    print("   ✅ États générés")
    
    # Ajouter les balances pour l'enrichissement
    results['balance_n_df'] = balance_n
    results['balance_n1_df'] = balance_n1
    
    # Afficher le contenu
    print("\n🔍 Contenu des résultats:")
    print(f"   - Clés: {list(results.keys())}")
    
    if 'bilan_actif' in results:
        bilan_actif = results['bilan_actif']
        print(f"   - Bilan actif: {type(bilan_actif)} avec {len(bilan_actif)} éléments")
        
        if isinstance(bilan_actif, list) and len(bilan_actif) > 0:
            print("\n   Premiers postes:")
            for poste in bilan_actif[:3]:
                print(f"      {poste.get('ref')}: {poste.get('libelle', 'N/A')[:40]}")
    
    if 'tft' in results:
        tft = results['tft']
        print(f"   - TFT: {type(tft)} avec {len(tft) if isinstance(tft, (dict, list)) else 'N/A'} éléments")
        
        if isinstance(tft, dict):
            print("\n   Premières clés TFT:")
            for i, (cle, valeur) in enumerate(list(tft.items())[:3]):
                print(f"      {cle}: {valeur:,.0f}")
    
    # Charger le template
    print("\n📂 Chargement du template...")
    template_path = "Liasse_officielle_revise.xlsx"
    if not os.path.exists(template_path):
        print(f"   ⚠️ Template non trouvé: {template_path}")
        sys.exit(1)
    
    wb = load_workbook(template_path)
    print(f"   ✅ Template chargé: {len(wb.sheetnames)} onglets")
    print(f"   Onglets: {wb.sheetnames[:5]}...")
    
    # Vérifier l'onglet ACTIF
    onglet_actif = next((name for name in wb.sheetnames if 'ACTIF' in name.upper() and 'PASSIF' not in name.upper() and name != 'BILAN'), None)
    if onglet_actif:
        print(f"\n✅ Onglet ACTIF trouvé: {onglet_actif}")
        ws = wb[onglet_actif]
        
        # Lire quelques cellules pour voir la structure
        print("   Structure de l'onglet:")
        for row in range(1, 6):
            values = [ws.cell(row, col).value for col in range(1, 10)]
            print(f"      Ligne {row}: {values}")
    
    # Vérifier l'onglet TFT
    onglet_tft = next((name for name in wb.sheetnames if 'TFT' in name.upper()), None)
    if onglet_tft:
        print(f"\n✅ Onglet TFT trouvé: {onglet_tft}")
    
    print("\n" + "="*80)
    print("✅ TEST TERMINÉ - Vérification de la structure OK")
    print("="*80)
    print("\nPour tester l'export complet, il faut installer FastAPI:")
    print("   pip install fastapi openpyxl pandas")
    
except Exception as e:
    print(f"\n❌ ERREUR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

import os
