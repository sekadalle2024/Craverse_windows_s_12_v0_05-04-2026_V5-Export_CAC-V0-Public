"""
Script de diagnostic pour l'export de la liasse officielle
Vérifie:
1. Présence du template
2. Structure des onglets
3. Format des données results
4. Mappings de cellules
"""

import openpyxl
import os
import json
from pathlib import Path

def diagnostic_template():
    """Diagnostic du fichier template"""
    print("=" * 80)
    print("DIAGNOSTIC TEMPLATE LIASSE OFFICIELLE")
    print("=" * 80)
    print()
    
    # 1. Vérifier l'existence du template
    print("1. Vérification du fichier template...")
    template_path = "Liasse_officielle_revise.xlsx"
    
    if not os.path.exists(template_path):
        print(f"   ❌ Fichier non trouvé: {template_path}")
        print(f"   📂 Répertoire actuel: {os.getcwd()}")
        print(f"   📄 Fichiers disponibles:")
        for f in os.listdir('.'):
            if f.endswith('.xlsx') or f.endswith('.xlsm'):
                print(f"      - {f}")
        return False
    
    print(f"   ✅ Fichier trouvé: {template_path}")
    print(f"   📊 Taille: {os.path.getsize(template_path) / 1024:.2f} KB")
    print()
    
    # 2. Charger le workbook
    print("2. Chargement du workbook...")
    try:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        print(f"   ✅ Workbook chargé")
        print(f"   📋 Nombre d'onglets: {len(wb.sheetnames)}")
        print()
    except Exception as e:
        print(f"   ❌ Erreur: {e}")
        return False
    
    # 3. Lister les onglets
    print("3. Liste des onglets:")
    onglets_importants = ['BILAN', 'Bilan', 'ACTIF', 'Actif', 'PASSIF', 'Passif', 
                          'RESULTAT', 'Résultat', 'Compte de résultat', 'CR']
    
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        marqueur = "⭐" if sheet_name in onglets_importants else "  "
        print(f"   {marqueur} {i:2d}. {sheet_name}")
    print()
    
    # 4. Analyser les onglets clés
    print("4. Analyse des onglets clés:")
    
    onglets_a_analyser = {
        'BILAN': 'Bilan consolidé',
        'Bilan': 'Bilan consolidé',
        'ACTIF': 'Détail actif',
        'Actif': 'Détail actif',
        'PASSIF': 'Détail passif',
        'Passif': 'Détail passif',
        'RESULTAT': 'Compte de résultat',
        'Résultat': 'Compte de résultat',
        'Compte de résultat': 'Compte de résultat',
        'CR': 'Compte de résultat'
    }
    
    onglets_trouves = {}
    
    for onglet_nom, description in onglets_a_analyser.items():
        if onglet_nom in wb.sheetnames:
            ws = wb[onglet_nom]
            print(f"   ✅ {onglet_nom} ({description})")
            print(f"      - Dimensions: {ws.max_row} lignes x {ws.max_column} colonnes")
            
            # Lire quelques cellules pour voir la structure
            print(f"      - Échantillon de cellules:")
            for row in range(1, min(6, ws.max_row + 1)):
                for col in range(1, min(6, ws.max_column + 1)):
                    cell = ws.cell(row, col)
                    if cell.value:
                        print(f"         {cell.coordinate}: {str(cell.value)[:50]}")
            
            onglets_trouves[onglet_nom] = description
            print()
    
    if not onglets_trouves:
        print("   ⚠️ Aucun onglet clé trouvé!")
        print("   💡 Les noms d'onglets dans le template sont peut-être différents")
        print()
    
    # 5. Recommandations
    print("5. Recommandations:")
    
    if 'BILAN' not in wb.sheetnames and 'Bilan' not in wb.sheetnames:
        print("   ⚠️ Onglet BILAN non trouvé")
        print("      → Vérifier le nom exact de l'onglet bilan")
    
    if 'RESULTAT' not in wb.sheetnames and 'Résultat' not in wb.sheetnames and 'Compte de résultat' not in wb.sheetnames:
        print("   ⚠️ Onglet RESULTAT non trouvé")
        print("      → Vérifier le nom exact de l'onglet compte de résultat")
    
    print()
    print("=" * 80)
    print("DIAGNOSTIC TERMINÉ")
    print("=" * 80)
    
    wb.close()
    return True


def diagnostic_donnees_test():
    """Crée des données de test pour vérifier le remplissage"""
    print()
    print("=" * 80)
    print("GÉNÉRATION DONNÉES DE TEST")
    print("=" * 80)
    print()
    
    # Créer des données de test au format attendu
    results_test = {
        'bilan_actif': {
            'AD': {'ref': 'AD', 'libelle': 'Charges immobilisées', 'montant': 1000000},
            'AE': {'ref': 'AE', 'libelle': 'Frais de recherche', 'montant': 500000},
            'AI': {'ref': 'AI', 'libelle': 'Terrains', 'montant': 5000000},
            'AJ': {'ref': 'AJ', 'libelle': 'Bâtiments', 'montant': 10000000},
            'BB': {'ref': 'BB', 'libelle': 'Stocks', 'montant': 2000000},
            'BI': {'ref': 'BI', 'libelle': 'Clients', 'montant': 3000000},
            'BV': {'ref': 'BV', 'libelle': 'Banques', 'montant': 1500000},
        },
        'bilan_passif': {
            'DA': {'ref': 'DA', 'libelle': 'Capital', 'montant': 10000000},
            'DH': {'ref': 'DH', 'libelle': 'Résultat net', 'montant': 2000000},
            'RA': {'ref': 'RA', 'libelle': 'Emprunts', 'montant': 5000000},
            'TC': {'ref': 'TC', 'libelle': 'Fournisseurs', 'montant': 3000000},
            'TD': {'ref': 'TD', 'libelle': 'Dettes fiscales', 'montant': 1000000},
        },
        'charges': {
            'TA': {'ref': 'TA', 'libelle': 'Achats marchandises', 'montant': 5000000},
            'TK': {'ref': 'TK', 'libelle': 'Charges personnel', 'montant': 3000000},
            'TL': {'ref': 'TL', 'libelle': 'Dotations amortissements', 'montant': 1000000},
        },
        'produits': {
            'RA': {'ref': 'RA', 'libelle': 'Ventes marchandises', 'montant': 15000000},
            'RB': {'ref': 'RB', 'libelle': 'Ventes produits', 'montant': 5000000},
        }
    }
    
    # Sauvegarder dans un fichier JSON
    output_file = "test_data_export_liasse.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results_test, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Données de test créées: {output_file}")
    print()
    print("Structure des données:")
    print(f"  - Bilan Actif: {len(results_test['bilan_actif'])} postes")
    print(f"  - Bilan Passif: {len(results_test['bilan_passif'])} postes")
    print(f"  - Charges: {len(results_test['charges'])} postes")
    print(f"  - Produits: {len(results_test['produits'])} postes")
    print()
    
    return results_test


def test_remplissage_simple():
    """Test de remplissage avec données fictives"""
    print("=" * 80)
    print("TEST REMPLISSAGE SIMPLE")
    print("=" * 80)
    print()
    
    template_path = "Liasse_officielle_revise.xlsx"
    
    if not os.path.exists(template_path):
        print(f"❌ Template non trouvé: {template_path}")
        return False
    
    print("1. Chargement du template...")
    try:
        wb = openpyxl.load_workbook(template_path)
        print("   ✅ Template chargé")
    except Exception as e:
        print(f"   ❌ Erreur: {e}")
        return False
    
    print()
    print("2. Test d'écriture dans les cellules...")
    
    # Essayer d'écrire dans différents onglets possibles
    onglets_test = ['BILAN', 'Bilan', 'ACTIF', 'Actif']
    
    for onglet_nom in onglets_test:
        if onglet_nom in wb.sheetnames:
            print(f"   📝 Test écriture dans '{onglet_nom}'...")
            ws = wb[onglet_nom]
            
            # Essayer d'écrire dans quelques cellules
            test_cells = ['C10', 'C11', 'C12', 'E10', 'E11', 'E12']
            
            for cell_addr in test_cells:
                try:
                    ws[cell_addr] = 123456.78
                    print(f"      ✅ {cell_addr} = 123456.78")
                except Exception as e:
                    print(f"      ❌ {cell_addr}: {e}")
    
    print()
    print("3. Sauvegarde du fichier test...")
    
    output_file = "test_liasse_remplie.xlsx"
    try:
        wb.save(output_file)
        print(f"   ✅ Fichier sauvegardé: {output_file}")
        print(f"   📊 Taille: {os.path.getsize(output_file) / 1024:.2f} KB")
        print()
        print("   💡 Ouvrez ce fichier pour vérifier si les valeurs sont présentes")
    except Exception as e:
        print(f"   ❌ Erreur: {e}")
        return False
    
    wb.close()
    return True


if __name__ == "__main__":
    # Changer vers le répertoire py_backend
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    
    print(f"📂 Répertoire de travail: {os.getcwd()}")
    print()
    
    # 1. Diagnostic du template
    diagnostic_template()
    
    # 2. Générer des données de test
    diagnostic_donnees_test()
    
    # 3. Test de remplissage simple
    test_remplissage_simple()
    
    print()
    print("=" * 80)
    print("DIAGNOSTIC COMPLET TERMINÉ")
    print("=" * 80)
    print()
    print("Fichiers générés:")
    print("  - test_data_export_liasse.json (données de test)")
    print("  - test_liasse_remplie.xlsx (test de remplissage)")
    print()
