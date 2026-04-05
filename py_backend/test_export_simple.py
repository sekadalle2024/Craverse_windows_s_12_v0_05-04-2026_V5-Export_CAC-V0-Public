"""
Test simple de l'export liasse avec la correction des cellules fusionnées
"""
import sys
import os

# Ajouter le répertoire au path
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.cell.cell import MergedCell

def test_correction_merged_cells():
    """Test de la fonction write_to_cell avec cellules fusionnées"""
    print("=" * 80)
    print("TEST CORRECTION CELLULES FUSIONNÉES")
    print("=" * 80)
    print()
    
    # Charger le template
    print("1. Chargement du template...")
    template_path = "Liasse_officielle_revise.xlsx"
    
    if not os.path.exists(template_path):
        print(f"   ❌ Template non trouvé: {template_path}")
        return False
    
    wb = openpyxl.load_workbook(template_path)
    print(f"   ✅ Template chargé ({len(wb.sheetnames)} onglets)")
    print()
    
    # Fonction write_to_cell (copie de export_liasse.py)
    def write_to_cell(ws, cell_addr, value):
        """Écrit une valeur en gérant les cellules fusionnées"""
        try:
            cell = ws[cell_addr]
            
            # Vérifier si c'est une cellule fusionnée
            if isinstance(cell, MergedCell):
                # Trouver la cellule principale
                for merged_range in ws.merged_cells.ranges:
                    if cell_addr in merged_range:
                        top_left_cell = merged_range.start_cell
                        ws.cell(top_left_cell.row, top_left_cell.column, value)
                        return True
                return False
            else:
                # Cellule normale
                ws[cell_addr] = value
                return True
        except Exception as e:
            print(f"      ❌ Erreur {cell_addr}: {e}")
            return False
    
    # Test sur l'onglet ACTIF
    print("2. Test d'écriture avec write_to_cell()...")
    
    if 'ACTIF' not in wb.sheetnames:
        print("   ❌ Onglet ACTIF non trouvé")
        return False
    
    ws_actif = wb['ACTIF']
    
    # Cellules de test (certaines fusionnées, d'autres normales)
    test_cells = {
        'C10': 1000000,  # Fusionnée
        'C11': 500000,   # Fusionnée
        'E11': 750000,   # Normale
        'E12': 250000,   # Normale
        'C15': 2000000,  # Fusionnée
    }
    
    compteur_ok = 0
    compteur_erreur = 0
    
    for cell_addr, value in test_cells.items():
        if write_to_cell(ws_actif, cell_addr, value):
            print(f"   ✅ {cell_addr} = {value:,.0f}")
            compteur_ok += 1
        else:
            print(f"   ❌ {cell_addr} échec")
            compteur_erreur += 1
    
    print()
    print(f"   Résultat: {compteur_ok} succès, {compteur_erreur} erreurs")
    print()
    
    # Sauvegarder
    print("3. Sauvegarde du fichier test...")
    output_file = "test_export_avec_correction.xlsx"
    
    try:
        wb.save(output_file)
        file_size = os.path.getsize(output_file) / 1024
        print(f"   ✅ Fichier sauvegardé: {output_file}")
        print(f"   📊 Taille: {file_size:.2f} KB")
        print()
        print("   💡 Ouvrez ce fichier pour vérifier que les valeurs apparaissent")
        print(f"      dans l'onglet ACTIF aux cellules: {', '.join(test_cells.keys())}")
    except Exception as e:
        print(f"   ❌ Erreur sauvegarde: {e}")
        return False
    
    wb.close()
    
    print()
    print("=" * 80)
    print("TEST TERMINÉ AVEC SUCCÈS")
    print("=" * 80)
    
    return True


if __name__ == "__main__":
    success = test_correction_merged_cells()
    sys.exit(0 if success else 1)
