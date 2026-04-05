# -*- coding: utf-8 -*-
"""
Test de l'export liasse avec la fonction write_to_cell() corrigée
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import os

print("=" * 80)
print("TEST EXPORT LIASSE AVEC CORRECTION write_to_cell()")
print("=" * 80)
print()

# Fonction write_to_cell() copiée depuis export_liasse.py
def write_to_cell(ws, cell_addr, value):
    """
    Écrit une valeur dans une cellule en gérant les cellules fusionnées
    
    Args:
        ws: Worksheet
        cell_addr: Adresse de la cellule (ex: 'C10')
        value: Valeur à écrire
    
    Returns:
        bool: True si succès, False sinon
    """
    try:
        cell = ws[cell_addr]
        
        # Vérifier si c'est une cellule fusionnée
        if isinstance(cell, MergedCell):
            # Trouver la cellule principale de la fusion
            for merged_range in ws.merged_cells.ranges:
                if cell_addr in merged_range:
                    # Écrire dans la cellule en haut à gauche de la fusion
                    top_left_cell = merged_range.start_cell
                    ws.cell(top_left_cell.row, top_left_cell.column, value)
                    print(f"   OK {cell_addr} (fusionnee) = {value}")
                    return True
            print(f"   ERREUR {cell_addr}: Cellule fusionnee mais range non trouve")
            return False
        else:
            # Cellule normale
            ws[cell_addr] = value
            print(f"   OK {cell_addr} (normale) = {value}")
            return True
    except Exception as e:
        print(f"   ERREUR {cell_addr}: {e}")
        return False

# Charger le template
template_path = "Liasse_officielle_revise.xlsx"
if not os.path.exists(template_path):
    print(f"ERREUR: Template non trouve: {template_path}")
    exit(1)

print(f"1. Chargement du template: {template_path}")
wb = load_workbook(template_path)
print(f"   OK Template charge")
print()

# Test sur l'onglet ACTIF
if 'ACTIF' in wb.sheetnames:
    ws_actif = wb['ACTIF']
    print("2. Test ecriture dans ACTIF avec write_to_cell()...")
    
    # Tester les cellules qui étaient problématiques
    cellules_test = {
        'C10': 111111.11,
        'C11': 222222.22,
        'C12': 333333.33,
        'E10': 444444.44,
        'E11': 555555.55,
        'E12': 666666.66,
        'C15': 777777.77,  # Cellule normale pour comparaison
    }
    
    compteur_ok = 0
    compteur_erreur = 0
    
    for cellule, valeur in cellules_test.items():
        if write_to_cell(ws_actif, cellule, valeur):
            compteur_ok += 1
        else:
            compteur_erreur += 1
    
    print()
    print(f"   Resultat: {compteur_ok} OK, {compteur_erreur} ERREUR")
    print()
else:
    print("ERREUR: Onglet ACTIF non trouve")
    exit(1)

# Test sur l'onglet PASSIF
if 'PASSIF' in wb.sheetnames:
    ws_passif = wb['PASSIF']
    print("3. Test ecriture dans PASSIF avec write_to_cell()...")
    
    cellules_test = {
        'E10': 888888.88,
        'E11': 999999.99,
        'E12': 101010.10,
    }
    
    compteur_ok = 0
    compteur_erreur = 0
    
    for cellule, valeur in cellules_test.items():
        if write_to_cell(ws_passif, cellule, valeur):
            compteur_ok += 1
        else:
            compteur_erreur += 1
    
    print()
    print(f"   Resultat: {compteur_ok} OK, {compteur_erreur} ERREUR")
    print()

# Sauvegarder le fichier
output_file = "test_export_avec_correction.xlsx"
print(f"4. Sauvegarde du fichier: {output_file}")
wb.save(output_file)
file_size = os.path.getsize(output_file) / 1024
print(f"   OK Fichier sauvegarde ({file_size:.2f} KB)")
print()

print("=" * 80)
print("TEST TERMINE")
print("=" * 80)
print()
print("VERIFICATION:")
print(f"  1. Ouvrir le fichier: {output_file}")
print("  2. Aller dans l'onglet ACTIF")
print("  3. Verifier que les cellules C10, C11, C12, E10, E11, E12, C15 contiennent des valeurs")
print("  4. Aller dans l'onglet PASSIF")
print("  5. Verifier que les cellules E10, E11, E12 contiennent des valeurs")
print()
