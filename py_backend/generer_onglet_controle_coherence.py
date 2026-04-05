# -*- coding: utf-8 -*-
"""
Module pour générer l'onglet "Contrôle de cohérence" dans la liasse officielle
Contient les 16 états de contrôle organisés par sections
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
import logging

logger = logging.getLogger("controle_coherence")


def format_montant_excel(montant: float) -> str:
    """Formate un montant pour Excel"""
    if abs(montant) < 0.01:
        return "-"
    return f"{montant:,.0f}".replace(',', ' ')


def ajouter_onglet_controle_coherence(wb: Workbook, etats_controle: List[Dict[str, Any]]) -> None:
    """
    Ajoute un onglet "Contrôle de cohérence" au workbook avec les 16 états de contrôle
    
    Args:
        wb: Workbook openpyxl
        etats_controle: Liste des 16 états de contrôle
    """
    logger.info("📊 Création de l'onglet 'Contrôle de cohérence'...")
    
    # Créer ou récupérer l'onglet
    if "Contrôle de cohérence" in wb.sheetnames:
        ws = wb["Contrôle de cohérence"]
        # Effacer le contenu existant
        wb.remove(ws)
    
    ws = wb.create_sheet("Contrôle de cohérence", 0)  # Insérer en première position
    
    # Styles
    style_titre_principal = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    fill_titre_principal = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    
    style_section = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    fill_section = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    style_sous_section = Font(name='Arial', size=11, bold=True, color='000000')
    fill_sous_section = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    style_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    
    style_normal = Font(name='Arial', size=10)
    style_montant = Font(name='Arial', size=10, bold=False)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Largeurs des colonnes
    ws.column_dimensions['A'].width = 8   # REF
    ws.column_dimensions['B'].width = 50  # LIBELLÉ
    ws.column_dimensions['C'].width = 18  # EXERCICE N
    ws.column_dimensions['D'].width = 18  # EXERCICE N-1
    
    # Ligne actuelle
    row = 1
    
    # TITRE PRINCIPAL
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "CONTRÔLE DE COHÉRENCE DES ÉTATS FINANCIERS"
    cell.font = style_titre_principal
    cell.fill = fill_titre_principal
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 2
    
    # Sous-titre
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "16 États de Contrôle - SYSCOHADA Révisé"
    cell.font = Font(name='Arial', size=11, italic=True, color='666666')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # Organisation des états par sections
    sections = {
        'SECTION 1 : CONTRÔLES BILAN ACTIF': [1, 2, 3],
        'SECTION 2 : CONTRÔLES BILAN PASSIF': [4, 5, 6],
        'SECTION 3 : CONTRÔLES COMPTE DE RÉSULTAT': [7, 8, 9],
        'SECTION 4 : CONTRÔLES TABLEAU DES FLUX DE TRÉSORERIE': [10, 11, 12],
        'SECTION 5 : CONTRÔLES SENS DES COMPTES': [13, 14],
        'SECTION 6 : CONTRÔLES ÉQUILIBRE BILAN': [15, 16]
    }
    
    # Parcourir les sections
    for section_titre, etats_indices in sections.items():
        # Titre de section
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = section_titre
        cell.font = style_section
        cell.fill = fill_section
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cell.border = border_thin
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Parcourir les états de cette section
        for etat_idx in etats_indices:
            if etat_idx <= len(etats_controle):
                etat = etats_controle[etat_idx - 1]
                
                # Sous-titre de l'état
                ws.merge_cells(f'A{row}:D{row}')
                cell = ws[f'A{row}']
                cell.value = etat.get('titre', f'État {etat_idx}')
                cell.font = style_sous_section
                cell.fill = fill_sous_section
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
                cell.border = border_thin
                ws.row_dimensions[row].height = 20
                row += 1
                
                # En-têtes des colonnes
                headers = ['REF', 'LIBELLÉ', 'EXERCICE N', 'EXERCICE N-1']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = header
                    cell.font = style_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                ws.row_dimensions[row].height = 18
                row += 1
                
                # Lignes de données
                postes = etat.get('postes', [])
                for poste in postes:
                    # REF
                    cell = ws.cell(row=row, column=1)
                    cell.value = poste.get('ref', '')
                    cell.font = style_normal
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                    
                    # LIBELLÉ
                    cell = ws.cell(row=row, column=2)
                    cell.value = poste.get('libelle', '')
                    cell.font = style_normal
                    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                    cell.border = border_thin
                    
                    # EXERCICE N
                    cell = ws.cell(row=row, column=3)
                    montant_n = poste.get('montant_n', 0)
                    cell.value = format_montant_excel(montant_n)
                    cell.font = style_montant
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border_thin
                    
                    # EXERCICE N-1
                    cell = ws.cell(row=row, column=4)
                    montant_n1 = poste.get('montant_n1', 0)
                    cell.value = format_montant_excel(montant_n1)
                    cell.font = style_montant
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border_thin
                    
                    row += 1
                
                # Ligne vide après chaque état
                row += 1
        
        # Ligne vide après chaque section
        row += 1
    
    # Pied de page
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "Note : Les montants sont exprimés en FCFA. Un tiret (-) indique un montant nul ou non applicable."
    cell.font = Font(name='Arial', size=9, italic=True, color='666666')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30
    
    logger.info(f"✅ Onglet 'Contrôle de cohérence' créé avec {len(etats_controle)} états de contrôle")


def generer_etats_controle_pour_export(results: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Génère les 16 états de contrôle à partir des résultats des états financiers
    
    Args:
        results: Résultats des états financiers (bilan_actif, bilan_passif, compte_resultat, tft)
    
    Returns:
        Liste des 16 états de contrôle
    """
    from etats_controle_exhaustifs import (
        calculer_etat_controle_bilan_actif_n,
        calculer_etat_controle_bilan_actif_n1,
        calculer_etat_controle_bilan_actif_variation,
        calculer_etat_controle_bilan_passif_n,
        calculer_etat_controle_bilan_passif_n1,
        calculer_etat_controle_bilan_passif_variation,
        calculer_etat_controle_compte_resultat_n,
        calculer_etat_controle_compte_resultat_n1,
        calculer_etat_controle_compte_resultat_variation,
        calculer_etat_controle_tft_n,
        calculer_etat_controle_tft_n1,
        calculer_etat_controle_tft_variation,
        calculer_etat_controle_sens_comptes_n,
        calculer_etat_controle_sens_comptes_n1,
        calculer_etat_equilibre_bilan_n,
        calculer_etat_equilibre_bilan_n1
    )
    
    # Extraire les données
    bilan_actif = results.get('bilan_actif', [])
    bilan_passif = results.get('bilan_passif', [])
    compte_resultat = results.get('compte_resultat', [])
    tft = results.get('tft', {})
    
    # Extraire le résultat net
    resultat_net_n = 0
    resultat_net_n1 = 0
    if compte_resultat:
        # Le résultat net est le dernier poste (XI)
        dernier_poste = compte_resultat[-1]
        resultat_net_n = dernier_poste.get('montant_n', 0)
        resultat_net_n1 = dernier_poste.get('montant_n1', 0)
    
    # Balances (si disponibles)
    balance_n = results.get('balance_n', [])
    balance_n1 = results.get('balance_n1', [])
    
    # Générer les 16 états
    etats_controle = []
    
    # Section 1 : Bilan Actif
    etats_controle.append(calculer_etat_controle_bilan_actif_n(bilan_actif))
    etats_controle.append(calculer_etat_controle_bilan_actif_n1(bilan_actif))
    etats_controle.append(calculer_etat_controle_bilan_actif_variation(bilan_actif, bilan_actif))
    
    # Section 2 : Bilan Passif
    etats_controle.append(calculer_etat_controle_bilan_passif_n(bilan_passif))
    etats_controle.append(calculer_etat_controle_bilan_passif_n1(bilan_passif))
    etats_controle.append(calculer_etat_controle_bilan_passif_variation(bilan_passif, bilan_passif))
    
    # Section 3 : Compte de Résultat
    etats_controle.append(calculer_etat_controle_compte_resultat_n(compte_resultat))
    etats_controle.append(calculer_etat_controle_compte_resultat_n1(compte_resultat))
    etats_controle.append(calculer_etat_controle_compte_resultat_variation(compte_resultat, compte_resultat))
    
    # Section 4 : TFT
    tft_list = []  # Convertir le dict TFT en liste si nécessaire
    if tft:
        # Créer une liste de postes à partir du dict TFT
        for key, value in tft.items():
            if key.startswith('Z') or key.startswith('F'):
                tft_list.append({'ref': key, 'montant_n': value, 'montant_n1': 0})
    
    etats_controle.append(calculer_etat_controle_tft_n(tft_list))
    etats_controle.append(calculer_etat_controle_tft_n1(tft_list))
    etats_controle.append(calculer_etat_controle_tft_variation(tft_list, tft_list))
    
    # Section 5 : Sens des Comptes
    if balance_n:
        etats_controle.append(calculer_etat_controle_sens_comptes_n(balance_n))
    else:
        etats_controle.append({
            'titre': '13. Etat de contrôle Sens des Comptes (Exercice N)',
            'postes': [{'ref': 'SA', 'libelle': 'Données non disponibles', 'montant_n': 0, 'montant_n1': 0}]
        })
    
    if balance_n1:
        etats_controle.append(calculer_etat_controle_sens_comptes_n1(balance_n1))
    else:
        etats_controle.append({
            'titre': '14. Etat de contrôle Sens des Comptes (Exercice N-1)',
            'postes': [{'ref': 'SJ', 'libelle': 'Données non disponibles', 'montant_n': 0, 'montant_n1': 0}]
        })
    
    # Section 6 : Équilibre Bilan
    etats_controle.append(calculer_etat_equilibre_bilan_n(bilan_actif, bilan_passif, resultat_net_n))
    etats_controle.append(calculer_etat_equilibre_bilan_n1(bilan_actif, bilan_passif, resultat_net_n1))
    
    return etats_controle
