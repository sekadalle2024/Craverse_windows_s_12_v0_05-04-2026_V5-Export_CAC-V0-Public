# -*- coding: utf-8 -*-
"""
Vérification des cellules du template pour identifier les bonnes adresses
"""
import openpyxl
from openpyxl import load_workbook

print("=" * 80)
print("VERIFICATION CELLULES TEMPLATE")
print("=" * 80)
print()

# Charger le template
wb = load_workbook("Liasse_officielle_revise.xlsx")

# Vérifier l'onglet ACTIF
if 'ACTIF' in wb.sheetnames:
    ws = wb['ACTIF']
    print("ONGLET: ACTIF")
    print("-" * 80)
    
    # Chercher les libellés des postes pour identifier les bonnes lignes
    print("\nRecherche des libelles de postes...")
    print()
    
    for row in range(1, 50):  # Parcourir les 50 premières lignes
        cell_a = ws[f'A{row}'].value
        cell_b = ws[f'B{row}'].value
        
        if cell_a and isinstance(cell_a, str):
            # Chercher les mots-clés des postes
            keywords = ['immobilis', 'charge', 'frais', 'brevet', 'terrain', 
                       'batiment', 'materiel', 'stock', 'client', 'tresor']
            
            for keyword in keywords:
                if keyword.lower() in cell_a.lower():
                    print(f"Ligne {row:2d}: A={cell_a[:50]}")
                    print(f"          B={cell_b}")
                    print(f"          C={ws[f'C{row}'].value}")
                    print(f"          D={ws[f'D{row}'].value}")
                    print(f"          E={ws[f'E{row}'].value}")
                    print()
                    break
    
    print("\n" + "=" * 80)
    print("STRUCTURE DES COLONNES")
    print("=" * 80)
    print()
    
    # Afficher l'en-tête pour comprendre la structure
    print("En-tete (lignes 1-10):")
    for row in range(1, 11):
        print(f"Ligne {row:2d}:")
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            val = ws[f'{col}{row}'].value
            if val:
                print(f"  {col}{row}: {str(val)[:40]}")
        print()
    
    print("\n" + "=" * 80)
    print("CELLULES FUSIONNEES")
    print("=" * 80)
    print()
    
    # Lister les cellules fusionnées dans la zone d'intérêt
    print("Cellules fusionnees dans la zone C10:E30:")
    for merged_range in ws.merged_cells.ranges:
        # Vérifier si la fusion est dans notre zone d'intérêt
        if (merged_range.min_row >= 10 and merged_range.max_row <= 30 and
            merged_range.min_col >= 3 and merged_range.max_col <= 5):  # C=3, E=5
            print(f"  {merged_range}")
    
    print()

print("\n" + "=" * 80)
print("RECOMMANDATIONS")
print("=" * 80)
print()
print("1. Identifier les lignes exactes des postes dans le template")
print("2. Mettre a jour les mappings dans export_liasse.py")
print("3. Verifier que les colonnes C, D, E correspondent bien aux montants")
print()
