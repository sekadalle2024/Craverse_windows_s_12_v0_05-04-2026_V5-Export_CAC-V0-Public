# -*- coding: utf-8 -*-
"""
Module d'export de la liasse officielle Excel remplie avec les valeurs calculées
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import os
import shutil
from pathlib import Path
from datetime import datetime
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Dict, Any, Optional
import logging
import base64
import io

# Import du module de génération de l'onglet contrôle de cohérence
from generer_onglet_controle_coherence import (
    ajouter_onglet_controle_coherence,
    generer_etats_controle_pour_export
)

logger = logging.getLogger("export_liasse")

router = APIRouter(prefix="/export-liasse", tags=["Export Liasse"])


# ==================== MODÈLES PYDANTIC ====================

class ExportLiasseRequest(BaseModel):
    """Requête pour exporter la liasse officielle"""
    results: Dict[str, Any]  # Résultats des états financiers
    nom_entreprise: Optional[str] = "ENTREPRISE"
    exercice: Optional[str] = None

class ExportLiasseResponse(BaseModel):
    success: bool
    message: str
    file_base64: Optional[str] = None
    filename: Optional[str] = None


# ==================== MAPPING POSTES VERS CELLULES ====================

# Mapping des références de postes vers les cellules Excel de la liasse officielle
# Format: {'ref_poste': 'cellule_excel'}

MAPPING_BILAN_ACTIF = {
    # ACTIF IMMOBILISÉ
    'AD': 'C10',   # Charges immobilisées
    'AE': 'C11',   # Frais de recherche et développement
    'AF': 'C12',   # Brevets, licences, logiciels
    'AG': 'C13',   # Fonds commercial
    'AH': 'C14',   # Autres immobilisations incorporelles
    'AI': 'C15',   # Terrains
    'AJ': 'C16',   # Bâtiments
    'AK': 'C17',   # Installations et agencements
    'AL': 'C18',   # Matériel
    'AM': 'C19',   # Matériel de transport
    'AN': 'C20',   # Avances et acomptes versés sur immobilisations
    'AP': 'C21',   # Titres de participation
    'AQ': 'C22',   # Autres immobilisations financières
    'AZ': 'C23',   # TOTAL ACTIF IMMOBILISÉ
    
    # ACTIF CIRCULANT
    'BA': 'C25',   # Actif circulant HAO
    'BB': 'C26',   # Stocks et encours
    'BC': 'C27',   # Marchandises
    'BD': 'C28',   # Matières premières
    'BE': 'C29',   # Autres approvisionnements
    'BF': 'C30',   # En-cours
    'BG': 'C31',   # Produits fabriqués
    'BH': 'C32',   # Fournisseurs, avances versées
    'BI': 'C33',   # Clients
    'BJ': 'C34',   # Autres créances
    'BK': 'C35',   # Capital souscrit, appelé non versé
    'BQ': 'C36',   # TOTAL ACTIF CIRCULANT
    
    # TRÉSORERIE-ACTIF
    'BT': 'C38',   # Titres de placement
    'BU': 'C39',   # Valeurs à encaisser
    'BV': 'C40',   # Banques, chèques postaux, caisse
    'BZ': 'C41',   # TOTAL TRÉSORERIE-ACTIF
    
    # ÉCARTS DE CONVERSION-ACTIF
    'CA': 'C43',   # Diminution des créances
    'CB': 'C44',   # Augmentation des dettes
    'CZ': 'C45',   # TOTAL ÉCARTS DE CONVERSION-ACTIF
}

MAPPING_BILAN_PASSIF = {
    # CAPITAUX PROPRES
    'DA': 'E10',   # Capital
    'DB': 'E11',   # Apporteurs, capital non appelé
    'DC': 'E12',   # Primes liées au capital social
    'DD': 'E13',   # Écarts de réévaluation
    'DE': 'E14',   # Réserves indisponibles
    'DF': 'E15',   # Réserves libres
    'DG': 'E16',   # Report à nouveau
    'DH': 'E17',   # Résultat net de l'exercice
    'DI': 'E18',   # Subventions d'investissement
    'DJ': 'E19',   # Provisions réglementées
    'DZ': 'E20',   # TOTAL CAPITAUX PROPRES
    
    # DETTES FINANCIÈRES ET RESSOURCES ASSIMILÉES
    'RA': 'E22',   # Emprunts
    'RB': 'E23',   # Dettes de crédit-bail et contrats assimilés
    'RC': 'E24',   # Dettes financières diverses
    'RD': 'E25',   # Provisions financières pour risques et charges
    'RZ': 'E26',   # TOTAL DETTES FINANCIÈRES
    
    # PASSIF CIRCULANT
    'TA': 'E28',   # Dettes circulantes HAO
    'TB': 'E29',   # Clients, avances reçues
    'TC': 'E30',   # Fournisseurs d'exploitation
    'TD': 'E31',   # Dettes fiscales
    'TE': 'E32',   # Dettes sociales
    'TF': 'E33',   # Autres dettes
    'TG': 'E34',   # Provisions pour risques à court terme
    'TZ': 'E35',   # TOTAL PASSIF CIRCULANT
    
    # TRÉSORERIE-PASSIF
    'UA': 'E37',   # Banques, crédits d'escompte
    'UB': 'E38',   # Banques, crédits de trésorerie
    'UC': 'E39',   # Banques, découverts
    'UZ': 'E40',   # TOTAL TRÉSORERIE-PASSIF
    
    # ÉCARTS DE CONVERSION-PASSIF
    'VA': 'E42',   # Augmentation des créances
    'VB': 'E43',   # Diminution des dettes
    'VZ': 'E44',   # TOTAL ÉCARTS DE CONVERSION-PASSIF
}

# Mapping unifié Compte de Résultat (onglet RESULTAT)
# Charges: colonne C  |  Produits: colonne E
# (col D = N-1 Charges, col F = N-1 Produits - à compléter si nécessaire)
MAPPING_RESULTAT_CHARGES_COL_N = {
    # Charges d'exploitation
    'TA': 'C10',   # Achats de marchandises
    'TB': 'C11',   # Variation de stocks de marchandises
    'TC': 'C12',   # Achats de matières premières
    'TD': 'C13',   # Variation de stocks de matières
    'TE': 'C14',   # Autres achats
    'TF': 'C15',   # Variation de stocks d'autres approvisionnements
    'TG': 'C16',   # Transports
    'TH': 'C17',   # Services extérieurs
    'TI': 'C18',   # Impôts et taxes
    'TJ': 'C19',   # Autres charges
    'TK': 'C20',   # Charges de personnel
    'TL': 'C21',   # Dotations aux amortissements
    'TM': 'C22',   # Dotations aux provisions
    'TZ': 'C23',   # TOTAL CHARGES D'EXPLOITATION
    # Charges financières
    'UA': 'C25',   # Frais financiers
    'UB': 'C26',   # Pertes de change
    'UC': 'C27',   # Dotations aux amortissements et provisions financières
    'UZ': 'C28',   # TOTAL CHARGES FINANCIÈRES
    # Charges HAO
    'XA': 'C30',   # Valeurs comptables des cessions d'immobilisations
    'XB': 'C31',   # Charges HAO constatées
    'XC': 'C32',   # Dotations HAO
    'XZ': 'C33',   # TOTAL CHARGES HAO
}

MAPPING_RESULTAT_PRODUITS_COL_N = {
    # Produits d'exploitation
    'RA': 'E10',   # Ventes de marchandises
    'RB': 'E11',   # Ventes de produits fabriqués
    'RC': 'E12',   # Travaux, services vendus
    'RD': 'E13',   # Production stockée
    'RE': 'E14',   # Production immobilisée
    'RF': 'E15',   # Subventions d'exploitation
    'RG': 'E16',   # Autres produits
    'RH': 'E17',   # Reprises de provisions
    'RI': 'E18',   # Transferts de charges
    'RZ': 'E19',   # TOTAL PRODUITS D'EXPLOITATION
    # Produits financiers
    'SA': 'E21',   # Revenus financiers
    'SB': 'E22',   # Gains de change
    'SC': 'E23',   # Reprises de provisions financières
    'SD': 'E24',   # Transferts de charges financières
    'SZ': 'E25',   # TOTAL PRODUITS FINANCIERS
    # Produits HAO
    'YA': 'E27',   # Produits des cessions d'immobilisations
    'YB': 'E28',   # Produits HAO constatés
    'YC': 'E29',   # Reprises HAO
    'YD': 'E30',   # Transferts de charges HAO
    'YZ': 'E31',   # TOTAL PRODUITS HAO
}

MAPPING_COMPTE_RESULTAT_CHARGES = {
    # ACTIVITÉ D'EXPLOITATION
    'TA': 'C10',   # Achats de marchandises
    'TB': 'C11',   # Variation de stocks de marchandises
    'TC': 'C12',   # Achats de matières premières
    'TD': 'C13',   # Variation de stocks de matières
    'TE': 'C14',   # Autres achats
    'TF': 'C15',   # Variation de stocks d'autres approvisionnements
    'TG': 'C16',   # Transports
    'TH': 'C17',   # Services extérieurs
    'TI': 'C18',   # Impôts et taxes
    'TJ': 'C19',   # Autres charges
    'TK': 'C20',   # Charges de personnel
    'TL': 'C21',   # Dotations aux amortissements
    'TM': 'C22',   # Dotations aux provisions
    'TZ': 'C23',   # TOTAL CHARGES D'EXPLOITATION
    
    # ACTIVITÉ FINANCIÈRE
    'UA': 'C25',   # Frais financiers
    'UB': 'C26',   # Pertes de change
    'UC': 'C27',   # Dotations aux amortissements et provisions
    'UZ': 'C28',   # TOTAL CHARGES FINANCIÈRES
    
    # HORS ACTIVITÉS ORDINAIRES (HAO)
    'XA': 'C30',   # Valeurs comptables des cessions d'immobilisations
    'XB': 'C31',   # Charges HAO constatées
    'XC': 'C32',   # Dotations HAO
    'XZ': 'C33',   # TOTAL CHARGES HAO
}

MAPPING_COMPTE_RESULTAT_PRODUITS = {
    # ACTIVITÉ D'EXPLOITATION
    'RA': 'E10',   # Ventes de marchandises
    'RB': 'E11',   # Ventes de produits fabriqués
    'RC': 'E12',   # Travaux, services vendus
    'RD': 'E13',   # Production stockée
    'RE': 'E14',   # Production immobilisée
    'RF': 'E15',   # Subventions d'exploitation
    'RG': 'E16',   # Autres produits
    'RH': 'E17',   # Reprises de provisions
    'RI': 'E18',   # Transferts de charges
    'RZ': 'E19',   # TOTAL PRODUITS D'EXPLOITATION
    
    # ACTIVITÉ FINANCIÈRE
    'SA': 'E21',   # Revenus financiers
    'SB': 'E22',   # Gains de change
    'SC': 'E23',   # Reprises de provisions
    'SD': 'E24',   # Transferts de charges
    'SZ': 'E25',   # TOTAL PRODUITS FINANCIERS
    
    # HORS ACTIVITÉS ORDINAIRES (HAO)
    'YA': 'E27',   # Produits des cessions d'immobilisations
    'YB': 'E28',   # Produits HAO constatés
    'YC': 'E29',   # Reprises HAO
    'YD': 'E30',   # Transferts de charges
    'YZ': 'E31',   # TOTAL PRODUITS HAO
}


# ==================== FONCTIONS D'EXPORT ====================

def remplir_liasse_officielle(results: Dict[str, Any], nom_entreprise: str, exercice: str) -> bytes:
    """
    Remplit la liasse officielle avec les valeurs calculées
    
    Args:
        results: Résultats des états financiers
        nom_entreprise: Nom de l'entreprise
        exercice: Exercice comptable (ex: "2024")
    
    Returns:
        Contenu du fichier Excel en bytes
    """
    logger.info("📊 Début du remplissage de la liasse officielle")
    
    # Chemin du template (fichier vierge) - UTILISER Liasse_officielle_revise.xlsx
    template_path = "Liasse_officielle_revise.xlsx"
    if not os.path.exists(template_path):
        # Essayer avec les anciens noms en fallback
        template_path = "LIASSE.xlsx"
        if not os.path.exists(template_path):
            template_path = "Liasse officielle.xlsm"
            if not os.path.exists(template_path):
                raise FileNotFoundError("Fichier template de liasse non trouvé (Liasse_officielle_revise.xlsx, LIASSE.xlsx ou Liasse officielle.xlsm)")
    
    logger.info(f"📂 Template trouvé: {template_path}")
    
    # Charger le workbook
    wb = load_workbook(template_path)
    
    # Log des onglets disponibles
    logger.info(f"📋 Onglets disponibles: {wb.sheetnames[:10]}...")  # Premiers 10 onglets
    
    # Log du format des données reçues
    logger.info(f"📊 Données reçues - Clés: {list(results.keys())}")
    if 'bilan_actif' in results:
        logger.info(f"   Type bilan_actif: {type(results.get('bilan_actif'))}")
        if isinstance(results.get('bilan_actif'), list):
            logger.info(f"   Bilan actif (LIST): {len(results['bilan_actif'])} postes")
        elif isinstance(results.get('bilan_actif'), dict):
            logger.info(f"   Bilan actif (DICT): {len(results['bilan_actif'])} postes")
    
    # ==================== CONVERSION FORMAT DONNÉES ====================
    # Convertir LIST en DICT si nécessaire (etats_financiers_v2.py retourne des listes)
    
    def convert_list_to_dict(data):
        """Convertit une liste de postes en dictionnaire {ref: poste}"""
        if isinstance(data, list):
            return {poste['ref']: poste for poste in data}
        return data if isinstance(data, dict) else {}
    
    bilan_actif_dict = convert_list_to_dict(results.get('bilan_actif', {}))
    bilan_passif_dict = convert_list_to_dict(results.get('bilan_passif', {}))
    charges_dict = convert_list_to_dict(results.get('charges', {}))
    produits_dict = convert_list_to_dict(results.get('produits', {}))
    
    logger.info(f"📊 Données converties:")
    logger.info(f"   - Bilan Actif: {len(bilan_actif_dict)} postes")
    logger.info(f"   - Bilan Passif: {len(bilan_passif_dict)} postes")
    logger.info(f"   - Charges: {len(charges_dict)} postes")
    logger.info(f"   - Produits: {len(produits_dict)} postes")
    
    # ==================== ENRICHISSEMENT AVEC BRUT ET AMORTISSEMENT ====================
    # Récupérer les balances depuis results si disponibles
    balance_n_df = results.get('balance_n_df')
    balance_n1_df = results.get('balance_n1_df')
    
    if balance_n_df is not None:
        logger.info("📊 Enrichissement du Bilan ACTIF avec BRUT et AMORTISSEMENT...")
        enrichir_actif_avec_brut_amortissement(bilan_actif_dict, balance_n_df, balance_n1_df)
    else:
        logger.warning("⚠️ Balances non disponibles - BRUT et AMORTISSEMENT non calculés")
    
    # ==================== FONCTION POUR ÉCRIRE DANS CELLULE ====================
    
    def write_to_cell(ws, cell_addr, value):
        """
        Écrit une valeur dans une cellule en gérant les cellules fusionnées
        
        Args:
            ws: Worksheet
            cell_addr: Adresse de la cellule (ex: 'C10')
            value: Valeur à écrire
        
        Returns:
            bool: True si succès, False sinon
        """
        try:
            cell = ws[cell_addr]
            target_cell = cell
            
            # Vérifier si c'est une cellule fusionnée
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                # Trouver la cellule principale de la fusion
                for merged_range in ws.merged_cells.ranges:
                    if cell_addr in merged_range:
                        # Obtenir la vraie cellule de la fusion
                        target_cell = ws.cell(merged_range.start_cell.row, merged_range.start_cell.column)
                        break
            
            # [CRITIQUE] Nous écrasons TOUT, même les formules, pour assurer 
            # l'alignement strict avec les calculs du Backend (moteur SYSCOHADA).
            target_cell.value = value
            return True
        except Exception as e:
            logger.warning(f"   Erreur écriture {cell_addr}: {e}")
            return False
    
    def remplir_onglet_dynamique(ws, postes, col_debut, ligne_debut):
        """
        Remplit un onglet dynamiquement sans mapping fixe
        APPROCHE DYNAMIQUE: Lit les postes générés et les écrit séquentiellement
        
        Args:
            ws: Worksheet
            postes: Liste des postes [{ref, libelle, montant_n, montant_n1}, ...]
            col_debut: Colonne de début pour N (ex: 'C')
            ligne_debut: Ligne de début (ex: 10)
        
        Returns:
            int: Nombre de cellules remplies
        """
        compteur = 0
        ligne = ligne_debut
        
        # Colonne N-1 (suivante)
        col_n1 = chr(ord(col_debut) + 1)
        
        for poste in postes:
            # Écrire montant N
            montant_n = poste.get('montant_n', poste.get('montant', 0))
            if write_to_cell(ws, f'{col_debut}{ligne}', montant_n):
                compteur += 1
            
            # Écrire montant N-1
            montant_n1 = poste.get('montant_n1', 0)
            if write_to_cell(ws, f'{col_n1}{ligne}', montant_n1):
                compteur += 1
            
            ligne += 1
        
        return compteur
    
    # ==================== REMPLISSAGE DES ONGLETS ====================
    
    # Log des onglets disponibles pour diagnostic
    logger.info(f"📋 Onglets disponibles dans le template: {wb.sheetnames}")
    
    # Remplir les informations générales
    if 'BILAN' in wb.sheetnames:
        ws_bilan = wb['BILAN']
        # Cellules pour nom entreprise et exercice (à adapter selon le template)
        write_to_cell(ws_bilan, 'C3', nom_entreprise)
        write_to_cell(ws_bilan, 'E5', f"{exercice}")
    
    # ==================== NOUVELLE APPROCHE : SCANNER DE REF INTELLIGENT ====================
    logger.info("📝 Remplissage par SCANNER INTELLIGENT de la colonne REF...")
    
    total_cellules = 0
    erreurs_total = 0
    
    # ---- Helper interne ----
    def extraire_montant(poste, cle_n='montant_n', cle_fallback='montant'):
        """Extrait le montant N d'un poste, avec fallback"""
        v = poste.get(cle_n, poste.get(cle_fallback, 0))
        return 0 if v is None else float(v)
    
    def extraire_montant_n1(poste):
        """Extrait le montant N-1 d'un poste"""
        v = poste.get('montant_n1', 0)
        return 0 if v is None else float(v)
    
    def calculer_totalisations_actif(bilan_actif_dict):
        """Calcule les totalisations du bilan actif"""
        totalisations = {}
        
        # AZ: TOTAL ACTIF IMMOBILISÉ (AD à AQ)
        refs_immobilise = ['AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AP', 'AQ']
        total_n = sum(extraire_montant(bilan_actif_dict.get(r, {})) for r in refs_immobilise)
        total_n1 = sum(extraire_montant_n1(bilan_actif_dict.get(r, {})) for r in refs_immobilise)
        totalisations['AZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # BQ: TOTAL ACTIF CIRCULANT (BA à BK)
        refs_circulant = ['BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK']
        total_n = sum(extraire_montant(bilan_actif_dict.get(r, {})) for r in refs_circulant)
        total_n1 = sum(extraire_montant_n1(bilan_actif_dict.get(r, {})) for r in refs_circulant)
        totalisations['BQ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # BZ: TOTAL TRÉSORERIE-ACTIF (BT, BU, BV)
        refs_tresorerie = ['BT', 'BU', 'BV']
        total_n = sum(extraire_montant(bilan_actif_dict.get(r, {})) for r in refs_tresorerie)
        total_n1 = sum(extraire_montant_n1(bilan_actif_dict.get(r, {})) for r in refs_tresorerie)
        totalisations['BZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # CZ: TOTAL ÉCARTS DE CONVERSION-ACTIF (CA, CB)
        refs_ecarts = ['CA', 'CB']
        total_n = sum(extraire_montant(bilan_actif_dict.get(r, {})) for r in refs_ecarts)
        total_n1 = sum(extraire_montant_n1(bilan_actif_dict.get(r, {})) for r in refs_ecarts)
        totalisations['CZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # DZ: TOTAL GÉNÉRAL ACTIF (AZ + BQ + BZ + CZ)
        total_n = (totalisations['AZ']['montant_n'] + totalisations['BQ']['montant_n'] + 
                   totalisations['BZ']['montant_n'] + totalisations['CZ']['montant_n'])
        total_n1 = (totalisations['AZ']['montant_n1'] + totalisations['BQ']['montant_n1'] + 
                    totalisations['BZ']['montant_n1'] + totalisations['CZ']['montant_n1'])
        totalisations['DZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        logger.info(f"   ✅ Totalisations ACTIF calculées: AZ={totalisations['AZ']['montant_n']:,.0f}, DZ={totalisations['DZ']['montant_n']:,.0f}")
        return totalisations
    
    def extraire_brut_et_amortissement_depuis_balance(balance_df, compte_principal):
        """
        Extrait les valeurs brutes et amortissements pour un compte depuis la balance
        
        Args:
            balance_df: DataFrame de la balance
            compte_principal: Numéro de compte (ex: '21' pour immobilisations corporelles)
        
        Returns:
            tuple: (brut, amortissement, net)
        """
        if balance_df is None or balance_df.empty:
            return 0, 0, 0
        
        try:
            # Détecter les colonnes de la balance
            from etats_financiers import detect_balance_columns, clean_number
            col_map = detect_balance_columns(balance_df)
            
            # Valeur brute: solde débit des comptes 2xxx
            brut = 0
            for idx, row in balance_df.iterrows():
                numero = str(row.get(col_map['numero'], '')).strip()
                if numero.startswith(compte_principal):
                    solde_debit = clean_number(row.get(col_map['solde_debit'], 0)) if col_map['solde_debit'] else 0
                    brut += solde_debit
            
            # Amortissement: solde crédit des comptes 28xx correspondants
            compte_amort = '28' + compte_principal[1:] if len(compte_principal) >= 2 else '28'
            amortissement = 0
            for idx, row in balance_df.iterrows():
                numero = str(row.get(col_map['numero'], '')).strip()
                if numero.startswith(compte_amort):
                    solde_credit = clean_number(row.get(col_map['solde_credit'], 0)) if col_map['solde_credit'] else 0
                    amortissement += solde_credit
            
            # Net = Brut - Amortissement
            net = brut - amortissement
            
            return brut, amortissement, net
            
        except Exception as e:
            logger.warning(f"   Erreur extraction brut/amort pour {compte_principal}: {e}")
            return 0, 0, 0
    
    def enrichir_actif_avec_brut_amortissement(bilan_actif_dict, balance_n_df, balance_n1_df):
        """
        Enrichit le dictionnaire bilan_actif avec les colonnes brut et amortissement
        
        Cette fonction ajoute les clés 'brut_n', 'amort_n', 'brut_n1', 'amort_n1' 
        aux postes d'immobilisations (AD à AQ)
        """
        # Mapping des REF vers les comptes principaux
        MAPPING_REF_COMPTES = {
            'AD': '20',  # Charges immobilisées
            'AE': '201', # Frais de recherche et développement
            'AF': '21',  # Brevets, licences, logiciels (immobilisations incorporelles)
            'AG': '207', # Fonds commercial
            'AH': '20',  # Autres immobilisations incorporelles
            'AI': '22',  # Terrains
            'AJ': '23',  # Bâtiments
            'AK': '24',  # Installations et agencements
            'AL': '24',  # Matériel
            'AM': '245', # Matériel de transport
            'AN': '25',  # Avances et acomptes versés sur immobilisations
            'AP': '26',  # Titres de participation
            'AQ': '27',  # Autres immobilisations financières
        }
        
        logger.info("   📊 Enrichissement ACTIF avec BRUT et AMORTISSEMENT...")
        
        for ref, compte in MAPPING_REF_COMPTES.items():
            if ref in bilan_actif_dict:
                # Extraire pour N
                brut_n, amort_n, net_n = extraire_brut_et_amortissement_depuis_balance(balance_n_df, compte)
                
                # Extraire pour N-1
                brut_n1, amort_n1, net_n1 = extraire_brut_et_amortissement_depuis_balance(balance_n1_df, compte)
                
                # Ajouter au dictionnaire
                bilan_actif_dict[ref]['brut_n'] = brut_n
                bilan_actif_dict[ref]['amort_n'] = amort_n
                bilan_actif_dict[ref]['brut_n1'] = brut_n1
                bilan_actif_dict[ref]['amort_n1'] = amort_n1
                
                # Vérifier la cohérence: net devrait être proche de brut - amort
                net_calcule = brut_n - amort_n
                net_existant = extraire_montant(bilan_actif_dict[ref])
                
                if abs(net_calcule - net_existant) > 1:  # Tolérance de 1 pour les arrondis
                    logger.debug(f"      {ref}: Net calculé={net_calcule:,.0f} vs Net existant={net_existant:,.0f}")
        
        logger.info(f"   ✅ ACTIF enrichi avec BRUT et AMORTISSEMENT")
    
    def calculer_totalisations_actif(bilan_actif_dict):
        """Calcule les totalisations du bilan actif (avec brut et amortissement si disponibles)"""
        totalisations = {}
        
        # AZ: TOTAL ACTIF IMMOBILISÉ (AD à AQ)
        refs_immobilise = ['AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AP', 'AQ']
        total_n = sum(extraire_montant(bilan_actif_dict.get(r, {})) for r in refs_immobilise)
        total_n1 = sum(extraire_montant_n1(bilan_actif_dict.get(r, {})) for r in refs_immobilise)
        
        # Calculer aussi les totaux brut et amortissement si disponibles
        total_brut_n = sum(bilan_actif_dict.get(r, {}).get('brut_n', 0) for r in refs_immobilise)
        total_amort_n = sum(bilan_actif_dict.get(r, {}).get('amort_n', 0) for r in refs_immobilise)
        total_brut_n1 = sum(bilan_actif_dict.get(r, {}).get('brut_n1', 0) for r in refs_immobilise)
        total_amort_n1 = sum(bilan_actif_dict.get(r, {}).get('amort_n1', 0) for r in refs_immobilise)
        
        totalisations['AZ'] = {
            'montant_n': total_n, 
            'montant_n1': total_n1,
            'brut_n': total_brut_n,
            'amort_n': total_amort_n,
            'brut_n1': total_brut_n1,
            'amort_n1': total_amort_n1
        }
        
        # BQ: TOTAL ACTIF CIRCULANT (BA à BK)
        refs_circulant = ['BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK']
        total_n = sum(extraire_montant(bilan_actif_dict.get(r, {})) for r in refs_circulant)
        total_n1 = sum(extraire_montant_n1(bilan_actif_dict.get(r, {})) for r in refs_circulant)
        totalisations['BQ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # BZ: TOTAL TRÉSORERIE-ACTIF (BT, BU, BV)
        refs_tresorerie = ['BT', 'BU', 'BV']
        total_n = sum(extraire_montant(bilan_actif_dict.get(r, {})) for r in refs_tresorerie)
        total_n1 = sum(extraire_montant_n1(bilan_actif_dict.get(r, {})) for r in refs_tresorerie)
        totalisations['BZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # CZ: TOTAL ÉCARTS DE CONVERSION-ACTIF (CA, CB)
        refs_ecarts = ['CA', 'CB']
        total_n = sum(extraire_montant(bilan_actif_dict.get(r, {})) for r in refs_ecarts)
        total_n1 = sum(extraire_montant_n1(bilan_actif_dict.get(r, {})) for r in refs_ecarts)
        totalisations['CZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # DZ: TOTAL GÉNÉRAL ACTIF (AZ + BQ + BZ + CZ)
        total_n = (totalisations['AZ']['montant_n'] + totalisations['BQ']['montant_n'] + 
                   totalisations['BZ']['montant_n'] + totalisations['CZ']['montant_n'])
        total_n1 = (totalisations['AZ']['montant_n1'] + totalisations['BQ']['montant_n1'] + 
                    totalisations['BZ']['montant_n1'] + totalisations['CZ']['montant_n1'])
        
        # Totaux brut et amortissement pour DZ
        total_brut_n = totalisations['AZ'].get('brut_n', 0)
        total_amort_n = totalisations['AZ'].get('amort_n', 0)
        total_brut_n1 = totalisations['AZ'].get('brut_n1', 0)
        total_amort_n1 = totalisations['AZ'].get('amort_n1', 0)
        
        totalisations['DZ'] = {
            'montant_n': total_n, 
            'montant_n1': total_n1,
            'brut_n': total_brut_n,
            'amort_n': total_amort_n,
            'brut_n1': total_brut_n1,
            'amort_n1': total_amort_n1
        }
        
        logger.info(f"   ✅ Totalisations ACTIF calculées: AZ={totalisations['AZ']['montant_n']:,.0f}, DZ={totalisations['DZ']['montant_n']:,.0f}")
        return totalisations
    
    def calculer_totalisations_passif(bilan_passif_dict):
        """Calcule les totalisations du bilan passif"""
        totalisations = {}
        
        # DZ: TOTAL CAPITAUX PROPRES (DA à DJ)
        refs_capitaux = ['DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ']
        total_n = sum(extraire_montant(bilan_passif_dict.get(r, {})) for r in refs_capitaux)
        total_n1 = sum(extraire_montant_n1(bilan_passif_dict.get(r, {})) for r in refs_capitaux)
        totalisations['DZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # RZ: TOTAL DETTES FINANCIÈRES (RA à RD)
        refs_dettes_fin = ['RA', 'RB', 'RC', 'RD']
        total_n = sum(extraire_montant(bilan_passif_dict.get(r, {})) for r in refs_dettes_fin)
        total_n1 = sum(extraire_montant_n1(bilan_passif_dict.get(r, {})) for r in refs_dettes_fin)
        totalisations['RZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # TZ: TOTAL PASSIF CIRCULANT (TA à TG)
        refs_passif_circ = ['TA', 'TB', 'TC', 'TD', 'TE', 'TF', 'TG']
        total_n = sum(extraire_montant(bilan_passif_dict.get(r, {})) for r in refs_passif_circ)
        total_n1 = sum(extraire_montant_n1(bilan_passif_dict.get(r, {})) for r in refs_passif_circ)
        totalisations['TZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # UZ: TOTAL TRÉSORERIE-PASSIF (UA, UB, UC)
        refs_tresorerie = ['UA', 'UB', 'UC']
        total_n = sum(extraire_montant(bilan_passif_dict.get(r, {})) for r in refs_tresorerie)
        total_n1 = sum(extraire_montant_n1(bilan_passif_dict.get(r, {})) for r in refs_tresorerie)
        totalisations['UZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        # VZ: TOTAL ÉCARTS DE CONVERSION-PASSIF (VA, VB)
        refs_ecarts = ['VA', 'VB']
        total_n = sum(extraire_montant(bilan_passif_dict.get(r, {})) for r in refs_ecarts)
        total_n1 = sum(extraire_montant_n1(bilan_passif_dict.get(r, {})) for r in refs_ecarts)
        totalisations['VZ'] = {'montant_n': total_n, 'montant_n1': total_n1}
        
        logger.info(f"   ✅ Totalisations PASSIF calculées: DZ={totalisations['DZ']['montant_n']:,.0f}")
        return totalisations
    
    # Calculer et ajouter les totalisations
    logger.info("📊 Calcul des totalisations...")
    totalisations_actif = calculer_totalisations_actif(bilan_actif_dict)
    bilan_actif_dict.update(totalisations_actif)
    
    totalisations_passif = calculer_totalisations_passif(bilan_passif_dict)
    bilan_passif_dict.update(totalisations_passif)
        
    def remplir_onglet_par_scan(onglet_name, dict_donnees, col_n, col_n1, ref_col_idx=1):
        """Scanne la colonne (par defaut A=1) pour les REF et remplit directement col_n et col_n1."""
        if not dict_donnees:
            return 0, 0
            
        ws = wb[onglet_name]
        compteur = 0
        erreurs = 0
        
        for row in ws.iter_rows(min_col=ref_col_idx, max_col=ref_col_idx, min_row=5):
            cell = row[0]
            ref_val = str(cell.value or '').strip()
            
            if len(ref_val) == 2 and ref_val.isalpha() and ref_val.isupper() and ref_val in dict_donnees:
                row_num = cell.row
                poste = dict_donnees[ref_val]
                montant_n = extraire_montant(poste)
                montant_n1 = extraire_montant_n1(poste)
                
                # Écrire N
                if col_n:
                    if write_to_cell(ws, f"{col_n}{row_num}", montant_n):
                        compteur += 1
                    else:
                        erreurs += 1
                        
                # Écrire N-1
                if col_n1:
                    if write_to_cell(ws, f"{col_n1}{row_num}", montant_n1):
                        compteur += 1
                    else:
                        erreurs += 1
        
        return compteur, erreurs
    
    def remplir_onglet_actif_avec_brut_amort(onglet_name, dict_donnees, col_brut_n='F', col_amort_n='G', col_net_n='H', col_net_n1='I', ref_col_idx=1):
        """
        Remplit l'onglet ACTIF avec les 4 colonnes: BRUT N, AMORT N, NET N, NET N-1
        
        Args:
            onglet_name: Nom de l'onglet
            dict_donnees: Dictionnaire des postes avec brut_n, amort_n, montant_n, montant_n1
            col_brut_n: Colonne pour valeurs brutes N (défaut: F)
            col_amort_n: Colonne pour amortissements N (défaut: G)
            col_net_n: Colonne pour valeurs nettes N (défaut: H)
            col_net_n1: Colonne pour valeurs nettes N-1 (défaut: I)
            ref_col_idx: Index de la colonne REF (défaut: 1 = colonne A)
        
        Returns:
            tuple: (compteur cellules remplies, erreurs)
        """
        if not dict_donnees:
            return 0, 0
            
        ws = wb[onglet_name]
        compteur = 0
        erreurs = 0
        
        for row in ws.iter_rows(min_col=ref_col_idx, max_col=ref_col_idx, min_row=5):
            cell = row[0]
            ref_val = str(cell.value or '').strip()
            
            if len(ref_val) == 2 and ref_val.isalpha() and ref_val.isupper() and ref_val in dict_donnees:
                row_num = cell.row
                poste = dict_donnees[ref_val]
                
                # Extraire les valeurs
                brut_n = poste.get('brut_n', 0) or 0
                amort_n = poste.get('amort_n', 0) or 0
                net_n = extraire_montant(poste)
                net_n1 = extraire_montant_n1(poste)
                
                # Écrire BRUT N (colonne F)
                if col_brut_n and brut_n != 0:
                    if write_to_cell(ws, f"{col_brut_n}{row_num}", brut_n):
                        compteur += 1
                    else:
                        erreurs += 1
                
                # Écrire AMORT N (colonne G)
                if col_amort_n and amort_n != 0:
                    if write_to_cell(ws, f"{col_amort_n}{row_num}", amort_n):
                        compteur += 1
                    else:
                        erreurs += 1
                
                # Écrire NET N (colonne H)
                if col_net_n:
                    if write_to_cell(ws, f"{col_net_n}{row_num}", net_n):
                        compteur += 1
                    else:
                        erreurs += 1
                
                # Écrire NET N-1 (colonne I)
                if col_net_n1:
                    if write_to_cell(ws, f"{col_net_n1}{row_num}", net_n1):
                        compteur += 1
                    else:
                        erreurs += 1
        
        return compteur, erreurs

    # ---- BILAN (GLOBAL) ----
    if 'BILAN' in wb.sheetnames:
        logger.info(f"📝 Remplissage BILAN (Actif et Passif)...")
        # Actif dans BILAN: REF en A (1), N en H, N-1 en I
        c, e = remplir_onglet_par_scan('BILAN', bilan_actif_dict, 'H', 'I', ref_col_idx=1)
        total_cellules += c
        erreurs_total += e
        # Passif dans BILAN: REF en J (10), N en M, N-1 en N
        c, e = remplir_onglet_par_scan('BILAN', bilan_passif_dict, 'M', 'N', ref_col_idx=10)
        total_cellules += c
        erreurs_total += e

    # ---- BILAN ACTIF (ONGLET SÉPARÉ) ----
    onglet_actif = next((name for name in wb.sheetnames if 'ACTIF' in name.upper() and 'PASSIF' not in name.upper() and name != 'BILAN'), None)
    if onglet_actif:
        logger.info(f"📝 Remplissage {onglet_actif} avec BRUT, AMORTISSEMENT, NET...")
        # Vérifier si les données brut/amort sont disponibles
        has_brut_amort = any(poste.get('brut_n') is not None for poste in bilan_actif_dict.values())
        
        if has_brut_amort:
            # Utiliser la fonction spéciale avec 4 colonnes: F (brut), G (amort), H (net N), I (net N-1)
            compteur, erreurs = remplir_onglet_actif_avec_brut_amort(
                onglet_actif, bilan_actif_dict, 
                col_brut_n='F', col_amort_n='G', col_net_n='H', col_net_n1='I'
            )
            logger.info(f"   ✅ ACTIF avec BRUT/AMORT: {compteur} cellules remplies, {erreurs} erreurs")
        else:
            # Fallback: remplissage standard (colonnes H et I)
            logger.warning("   ⚠️ Données BRUT/AMORT non disponibles - remplissage standard")
            compteur, erreurs = remplir_onglet_par_scan(onglet_actif, bilan_actif_dict, 'H', 'I')
            logger.info(f"   ✅ ACTIF standard: {compteur} cellules remplies, {erreurs} erreurs")
        
        total_cellules += compteur
        erreurs_total += erreurs
    
    # ---- BILAN PASSIF (ONGLET SÉPARÉ) ----
    onglet_passif = next((name for name in wb.sheetnames if 'PASSIF' in name.upper() and name != 'BILAN'), None)
    if onglet_passif:
        logger.info(f"📝 Remplissage {onglet_passif}...")
        # Passif: N en H, N-1 en I
        compteur, erreurs = remplir_onglet_par_scan(onglet_passif, bilan_passif_dict, 'H', 'I')
        total_cellules += compteur
        erreurs_total += erreurs

    # ---- COMPTE DE RÉSULTAT ----
    # Construire le dict compte_resultat par ref pour accès rapide
    cr_dict = {}
    cr_raw = results.get('compte_resultat', [])
    if isinstance(cr_raw, dict):
        cr_dict = cr_raw
    else:
        cr_dict = {p['ref']: p for p in cr_raw if 'ref' in p}
        
    if not cr_dict:  # Fallback
        c_raw = results.get('charges', {})
        p_raw = results.get('produits', {})
        c_dict = {p['ref']: p for p in c_raw} if isinstance(c_raw, list) else c_raw
        p_dict = {p['ref']: p for p in p_raw} if isinstance(p_raw, list) else p_raw
        cr_dict = {**c_dict, **p_dict}

    onglet_resultat = next((name for name in wb.sheetnames if 'RESULTAT' in name.upper() or 'RÉSULTAT' in name.upper()), None)
    if onglet_resultat:
        logger.info(f"📝 Remplissage {onglet_resultat}...")
        # Compte de résultat: N en I, N-1 en J (Basé sur screenshot RESULTAT 2)
        compteur, erreurs = remplir_onglet_par_scan(onglet_resultat, cr_dict, 'I', 'J')
        total_cellules += compteur
        erreurs_total += erreurs
        
    # ---- TFT ----
    # Mapping complet des clés TFT vers REF
    MAPPING_TFT_COMPLET = {
        'ZA_tresorerie_ouverture': 'ZA',
        'ZB_flux_operationnels': 'ZB',
        'ZC_flux_investissement': 'ZC',
        'ZD_flux_capitaux_propres': 'ZD',
        'ZE_flux_capitaux_etrangers': 'ZE',
        'ZF_flux_financement': 'ZF',
        'ZG_variation_tresorerie': 'ZG',
        'ZH_tresorerie_cloture': 'ZH',
        'FA_cafg': 'FA',
        'FB_variation_actif_hao': 'FB',
        'FC_variation_stocks': 'FC',
        'FD_variation_creances': 'FD',
        'FE_variation_dettes': 'FE',
        'FF_decaissements_investissement': 'FF',
        'FG_encaissements_investissement': 'FG',
        'FH_augmentation_capital': 'FH',
        'FI_dividendes_verses': 'FI',
        'FJ_emprunts_nouveaux': 'FJ',
        'FK_remboursements_emprunts': 'FK',
    }
    
    tft_dict = {}
    tft_raw = results.get('tft', {})
    
    if isinstance(tft_raw, dict):
        logger.info(f"   TFT dict reçu avec {len(tft_raw)} clés: {list(tft_raw.keys())[:5]}...")
        
        for cle, montant in tft_raw.items():
            # Chercher dans le mapping complet
            if cle in MAPPING_TFT_COMPLET:
                ref = MAPPING_TFT_COMPLET[cle]
                tft_dict[ref] = {'montant_n': float(montant or 0), 'montant_n1': 0}
            else:
                # Fallback: extraire prefixe si existant ex FF_decaissement
                prefixe = cle.split('_')[0] if '_' in cle else cle[:2]
                if prefixe.isalpha() and prefixe.isupper() and len(prefixe) == 2:
                    tft_dict[prefixe] = {'montant_n': float(montant or 0), 'montant_n1': 0}
        
        logger.info(f"   TFT dict converti: {len(tft_dict)} postes (ex: {list(tft_dict.keys())[:5]})")
    elif isinstance(tft_raw, list):
        # Si c'est déjà une liste de postes
        tft_dict = {p['ref']: p for p in tft_raw if 'ref' in p}
        logger.info(f"   TFT liste reçue: {len(tft_dict)} postes")
                    
    onglet_tft = next((name for name in wb.sheetnames if 'TFT' in name.upper() or 'TRÉSORERIE' in name.upper() or 'TRESORERIE' in name.upper()), None)
    if onglet_tft:
        if tft_dict:
            logger.info(f"📝 Remplissage {onglet_tft} avec {len(tft_dict)} postes...")
            # TFT: N en I, N-1 en K (Basé sur screenshot TFT 1)
            compteur, erreurs = remplir_onglet_par_scan(onglet_tft, tft_dict, 'I', 'K')
            total_cellules += compteur
            erreurs_total += erreurs
            logger.info(f"   ✅ TFT: {compteur} cellules remplies, {erreurs} erreurs")
        else:
            logger.warning(f"   ⚠️ TFT dict vide - onglet non rempli")
    else:
        logger.warning("   ⚠️ Onglet TFT non trouvé dans le template")
        
    # ---- BILAN ETAT COMPLET (Col A actif, Col J passif) ----
    onglet_bilan = next((name for name in wb.sheetnames if name.strip() == 'BILAN'), None)
    if onglet_bilan:
        logger.info(f"📝 Remplissage {onglet_bilan}...")
        # Actif scan col A (idx=1), valeurs en H et I
        c_actif, err_actif = remplir_onglet_par_scan(onglet_bilan, bilan_actif_dict, 'H', 'I', ref_col_idx=1)
        # Passif scan col J (idx=10), valeurs en M et N
        c_passif, err_passif = remplir_onglet_par_scan(onglet_bilan, bilan_passif_dict, 'M', 'N', ref_col_idx=10)
        
        total_cellules += (c_actif + c_passif)
        erreurs_total += (err_actif + err_passif)

    logger.info(f"✅ TOTAL MAPPING: {total_cellules} cellules remplies, {erreurs_total} erreurs")
    logger.info("✅ Remplissage intelligent terminé")
    
    # ==================== AJOUT ONGLET CONTRÔLE DE COHÉRENCE ====================
    logger.info("📊 Ajout de l'onglet 'Contrôle de cohérence'...")
    try:
        # Générer les 16 états de contrôle
        etats_controle = generer_etats_controle_pour_export(results)
        
        # Ajouter l'onglet au workbook
        ajouter_onglet_controle_coherence(wb, etats_controle)
        
        logger.info("✅ Onglet 'Contrôle de cohérence' ajouté avec succès")
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'ajout de l'onglet Contrôle de cohérence: {e}", exc_info=True)
        logger.warning("⚠️ L'export continue sans l'onglet Contrôle de cohérence")
    
    # ==================== FIN AJOUT ONGLET ====================
    
    # Sauvegarder dans un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info("✅ Liasse officielle remplie avec succès")
    
    return output.getvalue()


# ==================== ENDPOINT API ====================

@router.post("/generer", response_model=ExportLiasseResponse)
async def generer_liasse(request: ExportLiasseRequest):
    """
    Génère la liasse officielle Excel remplie avec les valeurs calculées
    """
    try:
        logger.info("📥 Réception demande d'export liasse")
        
        # Déterminer l'exercice
        exercice = request.exercice or datetime.now().year
        
        # Remplir la liasse
        file_content = remplir_liasse_officielle(
            results=request.results,
            nom_entreprise=request.nom_entreprise,
            exercice=str(exercice)
        )
        
        # Encoder en base64
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        
        # Nom du fichier
        filename = f"Liasse_Officielle_{request.nom_entreprise}_{exercice}.xlsx"
        filename = filename.replace(' ', '_').replace('/', '_')
        
        logger.info(f"✅ Liasse générée: {filename}")
        
        return ExportLiasseResponse(
            success=True,
            message=f"Liasse officielle générée avec succès pour {request.nom_entreprise} - Exercice {exercice}",
            file_base64=file_base64,
            filename=filename
        )
        
    except FileNotFoundError as e:
        logger.error(f"❌ Fichier template non trouvé: {e}")
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"❌ Erreur lors de la génération: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
