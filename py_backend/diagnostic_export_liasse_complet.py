# -*- coding: utf-8 -*-
"""
Script de diagnostic complet pour l'export de la liasse officielle
Identifie les problèmes de remplissage des cellules Excel
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import sys
from pathlib import Path

# Ajouter le répertoire parent au path pour les imports
sys.path.insert(0, str(Path(__file__).parent))

# Import conditionnel - la fonction sera créée si elle n'existe pas
try:
    from etats_financiers_v2 import generer_etats_financiers
except ImportError:
    # Fonction temporaire pour le diagnostic
    def generer_etats_financiers(balance_n, balance_n1, balance_n2):
        """Fonction temporaire pour le diagnostic"""
        return {
            'bilan_actif': [],
            'bilan_passif': [],
            'compte_resultat': [],
            'tft': {}
        }

def diagnostic_complet():
    """Diagnostic complet de l'export de la liasse"""
    
    print("=" * 80)
    print("DIAGNOSTIC COMPLET - EXPORT LIASSE OFFICIELLE")
    print("=" * 80)
    print()
    
    # ==================== ÉTAPE 1: VÉRIFIER LES FICHIERS ====================
    print("📁 ÉTAPE 1: Vérification des fichiers")
    print("-" * 80)
    
    # Fichier balance
    balance_file = "P000 -BALANCE DEMO N_N-1_N-2.xls"
    if not os.path.exists(balance_file):
        print(f"❌ Fichier balance non trouvé: {balance_file}")
        return
    print(f"✅ Fichier balance trouvé: {balance_file}")
    
    # Fichier template liasse
    template_file = "Liasse_officielle_revise.xlsx"
    if not os.path.exists(template_file):
        print(f"❌ Fichier template non trouvé: {template_file}")
        return
    print(f"✅ Fichier template trouvé: {template_file}")
    print()
    
    # ==================== ÉTAPE 2: CHARGER LES BALANCES ====================
    print("📊 ÉTAPE 2: Chargement des balances")
    print("-" * 80)
    
    try:
        xls = pd.ExcelFile(balance_file)
        print(f"Onglets disponibles: {xls.sheet_names}")
        
        balance_n = pd.read_excel(balance_file, sheet_name="BALANCE N")
        balance_n1 = pd.read_excel(balance_file, sheet_name="BALANCE N-1")
        balance_n2 = pd.read_excel(balance_file, sheet_name="BALANCE N-2")
        
        print(f"✅ Balance N: {len(balance_n)} comptes")
        print(f"✅ Balance N-1: {len(balance_n1)} comptes")
        print(f"✅ Balance N-2: {len(balance_n2)} comptes")
        print()
    except Exception as e:
        print(f"❌ Erreur chargement balances: {e}")
        return
    
    # ==================== ÉTAPE 3: GÉNÉRER LES ÉTATS FINANCIERS ====================
    print("🔢 ÉTAPE 3: Génération des états financiers")
    print("-" * 80)
    
    try:
        results = generer_etats_financiers(balance_n, balance_n1, balance_n2)
        
        print(f"✅ États générés:")
        print(f"   - Bilan Actif: {len(results.get('bilan_actif', []))} postes")
        print(f"   - Bilan Passif: {len(results.get('bilan_passif', []))} postes")
        print(f"   - Compte Résultat: {len(results.get('compte_resultat', []))} postes")
        print(f"   - TFT: {len(results.get('tft', {}))} éléments")
        print()
        
        # Afficher quelques valeurs pour vérification
        print("📋 Échantillon Bilan Actif (premiers 5 postes):")
        bilan_actif = results.get('bilan_actif', [])
        for i, poste in enumerate(bilan_actif[:5]):
            ref = poste.get('ref', '?')
            libelle = poste.get('libelle', '?')[:40]
            montant_n = poste.get('montant_n', 0)
            montant_n1 = poste.get('montant_n1', 0)
            print(f"   {ref}: {libelle:40} | N: {montant_n:>15,.0f} | N-1: {montant_n1:>15,.0f}")
        print()
        
    except Exception as e:
        print(f"❌ Erreur génération états: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # ==================== ÉTAPE 4: ANALYSER LE TEMPLATE EXCEL ====================
    print("📄 ÉTAPE 4: Analyse du template Excel")
    print("-" * 80)
    
    try:
        wb = load_workbook(template_file)
        print(f"Onglets dans le template: {wb.sheetnames[:10]}...")
        print()
        
        # Analyser l'onglet BILAN ACTIF
        onglet_actif = next((name for name in wb.sheetnames if 'ACTIF' in name.upper() and 'PASSIF' not in name.upper() and name != 'BILAN'), None)
        
        if onglet_actif:
            print(f"📊 Analyse de l'onglet: {onglet_actif}")
            ws = wb[onglet_actif]
            
            # Scanner la colonne A pour trouver les REF
            print("\n🔍 Scan de la colonne A (REF):")
            refs_trouvees = []
            for row in ws.iter_rows(min_col=1, max_col=1, min_row=5, max_row=50):
                cell = row[0]
                ref_val = str(cell.value or '').strip()
                if len(ref_val) == 2 and ref_val.isalpha() and ref_val.isupper():
                    refs_trouvees.append((cell.row, ref_val))
            
            print(f"   Trouvé {len(refs_trouvees)} références:")
            for row_num, ref in refs_trouvees[:10]:
                print(f"      Ligne {row_num}: {ref}")
            print()
            
            # Analyser les cellules fusionnées
            print("🔗 Cellules fusionnées dans la zone de données:")
            merged_count = 0
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row >= 5 and merged_range.min_row <= 50:
                    merged_count += 1
                    if merged_count <= 5:
                        print(f"      {merged_range}")
            print(f"   Total: {merged_count} zones fusionnées")
            print()
            
            # Vérifier les colonnes H et I (N et N-1)
            print("📍 Vérification des colonnes H et I (valeurs N et N-1):")
            if refs_trouvees:
                row_test = refs_trouvees[0][0]
                cell_h = ws[f'H{row_test}']
                cell_i = ws[f'I{row_test}']
                print(f"   Ligne {row_test}:")
                print(f"      H{row_test}: {cell_h.value} (Type: {type(cell_h).__name__})")
                print(f"      I{row_test}: {cell_i.value} (Type: {type(cell_i).__name__})")
            print()
            
        # Analyser l'onglet BILAN (global)
        if 'BILAN' in wb.sheetnames:
            print(f"📊 Analyse de l'onglet: BILAN")
            ws_bilan = wb['BILAN']
            
            print("\n🔍 Scan colonne A (ACTIF):")
            refs_actif = []
            for row in ws_bilan.iter_rows(min_col=1, max_col=1, min_row=5, max_row=50):
                cell = row[0]
                ref_val = str(cell.value or '').strip()
                if len(ref_val) == 2 and ref_val.isalpha() and ref_val.isupper():
                    refs_actif.append((cell.row, ref_val))
            print(f"   Trouvé {len(refs_actif)} références ACTIF")
            
            print("\n🔍 Scan colonne J (PASSIF):")
            refs_passif = []
            for row in ws_bilan.iter_rows(min_col=10, max_col=10, min_row=5, max_row=50):
                cell = row[0]
                ref_val = str(cell.value or '').strip()
                if len(ref_val) == 2 and ref_val.isalpha() and ref_val.isupper():
                    refs_passif.append((cell.row, ref_val))
            print(f"   Trouvé {len(refs_passif)} références PASSIF")
            print()
            
    except Exception as e:
        print(f"❌ Erreur analyse template: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # ==================== ÉTAPE 5: TESTER L'ÉCRITURE ====================
    print("✍️ ÉTAPE 5: Test d'écriture dans le template")
    print("-" * 80)
    
    try:
        # Créer une copie de travail
        wb_test = load_workbook(template_file)
        
        if onglet_actif:
            ws_test = wb_test[onglet_actif]
            
            # Tester l'écriture sur la première REF trouvée
            if refs_trouvees:
                row_test = refs_trouvees[0][0]
                ref_test = refs_trouvees[0][1]
                
                print(f"Test d'écriture sur ligne {row_test} (REF: {ref_test}):")
                
                # Trouver le poste correspondant dans bilan_actif
                poste_test = next((p for p in bilan_actif if p.get('ref') == ref_test), None)
                
                if poste_test:
                    montant_n = poste_test.get('montant_n', 0)
                    montant_n1 = poste_test.get('montant_n1', 0)
                    
                    print(f"   Valeurs à écrire:")
                    print(f"      N: {montant_n:,.0f}")
                    print(f"      N-1: {montant_n1:,.0f}")
                    
                    # Écrire dans H et I
                    ws_test[f'H{row_test}'].value = montant_n
                    ws_test[f'I{row_test}'].value = montant_n1
                    
                    # Vérifier
                    val_h = ws_test[f'H{row_test}'].value
                    val_i = ws_test[f'I{row_test}'].value
                    
                    print(f"   Valeurs écrites:")
                    print(f"      H{row_test}: {val_h}")
                    print(f"      I{row_test}: {val_i}")
                    
                    if val_h == montant_n and val_i == montant_n1:
                        print("   ✅ Écriture réussie!")
                    else:
                        print("   ❌ Écriture échouée - valeurs différentes")
                else:
                    print(f"   ⚠️ Poste {ref_test} non trouvé dans bilan_actif")
                
                # Sauvegarder le test
                test_output = "test_export_diagnostic.xlsx"
                wb_test.save(test_output)
                print(f"\n✅ Fichier de test sauvegardé: {test_output}")
                print()
                
    except Exception as e:
        print(f"❌ Erreur test écriture: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # ==================== ÉTAPE 6: DIAGNOSTIC DES PROBLÈMES ====================
    print("🔍 ÉTAPE 6: Diagnostic des problèmes identifiés")
    print("-" * 80)
    
    print("\n📋 PROBLÈME 1: Bilan ACTIF - Colonnes Brut et Amortissement")
    print("   État actuel:")
    print("      - Le menu accordéon affiche seulement la colonne NET pour N")
    print("      - Le template Excel a des colonnes BRUT et AMORTISSEMENT")
    print("   Solution:")
    print("      1. Ajouter les colonnes BRUT et AMORTISSEMENT dans le menu accordéon")
    print("      2. Calculer ces valeurs depuis la balance (comptes 2xxx)")
    print("      3. Remplir les cellules correspondantes dans l'export")
    print()
    
    print("📋 PROBLÈME 2: Bilan ACTIF - Totalisation manquante")
    print("   État actuel:")
    print("      - Les lignes de totalisation ne sont pas renseignées")
    print("   Solution:")
    print("      1. Identifier les lignes de totalisation (AZ, BQ, BZ, etc.)")
    print("      2. Calculer les totaux depuis les postes")
    print("      3. Remplir les cellules de totalisation")
    print()
    
    print("📋 PROBLÈME 3: TFT vierge")
    print("   État actuel:")
    print("      - L'onglet TFT reste vide après export")
    print("   Solution:")
    print("      1. Vérifier que le TFT est bien généré (format dict)")
    print("      2. Convertir le dict en liste de postes avec REF")
    print("      3. Scanner la colonne A du TFT pour les REF")
    print("      4. Remplir les colonnes I (N) et K (N-1)")
    print()
    
    print("=" * 80)
    print("DIAGNOSTIC TERMINÉ")
    print("=" * 80)
    print()
    print("📝 Prochaines étapes:")
    print("   1. Examiner le fichier test_export_diagnostic.xlsx")
    print("   2. Vérifier que les valeurs sont bien écrites en H et I")
    print("   3. Appliquer les corrections identifiées")
    print()

if __name__ == "__main__":
    diagnostic_complet()
