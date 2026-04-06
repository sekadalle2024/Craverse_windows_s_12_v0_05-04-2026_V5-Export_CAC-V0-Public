"""
Test final des corrections des problèmes 1 et 3
Teste directement avec la balance de démo
"""
import sys
import os
sys.path.insert(0, 'py_backend')

import pandas as pd
from io import BytesIO

print("=" * 80)
print("TEST FINAL - CORRECTIONS PROBLÈMES 1 ET 3")
print("=" * 80)
print()

# Étape 1: Charger la balance
print("📂 Chargement de la balance de test...")
balance_file = "py_backend/P000 -BALANCE DEMO N_N-1_N-2.xls"

if not os.path.exists(balance_file):
    print(f"❌ ERREUR: Fichier balance non trouvé: {balance_file}")
    sys.exit(1)

try:
    # Attention aux espaces dans les noms d'onglets!
    balance_n = pd.read_excel(balance_file, sheet_name="BALANCE N ")
    balance_n1 = pd.read_excel(balance_file, sheet_name="BALANCE N-1 ")
    balance_n2 = pd.read_excel(balance_file, sheet_name="BALANCE N-2")
    print(f"✅ Balance chargée: {len(balance_n)} lignes (N), {len(balance_n1)} lignes (N-1)")
except Exception as e:
    print(f"❌ ERREUR lors du chargement: {e}")
    sys.exit(1)

print()

# Étape 2: Générer les états financiers
print("📊 Génération des états financiers...")
try:
    from etats_financiers_v2 import process_balance_to_liasse_format
    import json
    
    # Charger les correspondances
    correspondances_file = "py_backend/correspondances_syscohada.json"
    if os.path.exists(correspondances_file):
        with open(correspondances_file, 'r', encoding='utf-8') as f:
            correspondances = json.load(f)
        print(f"✅ Correspondances chargées")
    else:
        print(f"⚠️  Fichier correspondances non trouvé, utilisation d'un dict vide")
        correspondances = {}
    
    results = process_balance_to_liasse_format(balance_n, balance_n1, balance_n2, correspondances)
    print(f"✅ États générés")
    print(f"   Clés disponibles: {list(results.keys())}")
    
    # Vérifier les balances dans results
    if 'balance_n_df' in results:
        print(f"   ✅ balance_n_df présente")
    else:
        print(f"   ⚠️  balance_n_df ABSENTE - Problème 1 ne fonctionnera pas")
    
    if 'balance_n1_df' in results:
        print(f"   ✅ balance_n1_df présente")
    else:
        print(f"   ⚠️  balance_n1_df ABSENTE - Problème 1 ne fonctionnera pas")
    
    # Vérifier le TFT
    if 'tft' in results:
        tft = results['tft']
        if isinstance(tft, dict):
            print(f"   ✅ TFT présent (dict avec {len(tft)} clés)")
            # Afficher quelques clés
            cles_exemple = list(tft.keys())[:3]
            print(f"      Exemples de clés: {cles_exemple}")
        elif isinstance(tft, list):
            print(f"   ✅ TFT présent (list avec {len(tft)} postes)")
        else:
            print(f"   ⚠️  TFT présent mais format inconnu: {type(tft)}")
    else:
        print(f"   ⚠️  TFT ABSENT - Problème 3 ne fonctionnera pas")
    
except Exception as e:
    print(f"❌ ERREUR lors de la génération: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print()

# Étape 3: Ajouter les balances aux results si absentes
print("🔧 Vérification et ajout des balances aux results...")
if 'balance_n_df' not in results:
    print("   ⚠️  Ajout manuel de balance_n_df")
    results['balance_n_df'] = balance_n
if 'balance_n1_df' not in results:
    print("   ⚠️  Ajout manuel de balance_n1_df")
    results['balance_n1_df'] = balance_n1

print()

# Étape 4: Exporter la liasse
print("📝 Export de la liasse officielle...")
try:
    from export_liasse import remplir_liasse_officielle
    
    file_content = remplir_liasse_officielle(
        results=results,
        nom_entreprise="TEST ENTREPRISE",
        exercice="2024"
    )
    
    print(f"✅ Export réussi: {len(file_content)} bytes")
    
    # Sauvegarder le fichier
    output_file = "test_corrections_final.xlsx"
    with open(output_file, "wb") as f:
        f.write(file_content)
    
    print(f"✅ Fichier sauvegardé: {output_file}")
    
except Exception as e:
    print(f"❌ ERREUR lors de l'export: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print()

# Étape 5: Vérifier le contenu du fichier exporté
print("🔍 Vérification du fichier exporté...")
try:
    from openpyxl import load_workbook
    
    wb = load_workbook(output_file)
    print(f"✅ Fichier Excel valide")
    print(f"   Onglets: {wb.sheetnames[:10]}")
    
    # Vérifier l'onglet ACTIF
    if 'ACTIF' in wb.sheetnames:
        ws_actif = wb['ACTIF']
        print()
        print("📋 Vérification onglet ACTIF:")
        
        # Chercher une ligne avec REF (ex: AI pour Terrains)
        ref_trouve = False
        for row in ws_actif.iter_rows(min_row=5, max_row=50, min_col=1, max_col=1):
            cell = row[0]
            if cell.value == 'AI':  # Terrains
                row_num = cell.row
                ref_trouve = True
                
                # Lire les colonnes E, F, G, H
                brut_n = ws_actif[f'E{row_num}'].value
                amort_n = ws_actif[f'F{row_num}'].value
                net_n = ws_actif[f'G{row_num}'].value
                net_n1 = ws_actif[f'H{row_num}'].value
                
                print(f"   Ligne AI (Terrains) - Ligne {row_num}:")
                print(f"      Colonne E (BRUT N): {brut_n}")
                print(f"      Colonne F (AMORT N): {amort_n}")
                print(f"      Colonne G (NET N): {net_n}")
                print(f"      Colonne H (NET N-1): {net_n1}")
                
                if brut_n is not None and brut_n != 0:
                    print(f"   ✅ PROBLÈME 1 RÉSOLU: Colonnes BRUT/AMORT remplies!")
                else:
                    print(f"   ⚠️  PROBLÈME 1 PERSISTE: Colonnes BRUT/AMORT vides")
                
                break
        
        if not ref_trouve:
            print(f"   ⚠️  REF 'AI' non trouvée dans l'onglet ACTIF")
    
    # Vérifier l'onglet TFT
    onglet_tft = None
    for name in wb.sheetnames:
        if 'TFT' in name.upper():
            onglet_tft = name
            break
    
    if onglet_tft:
        ws_tft = wb[onglet_tft]
        print()
        print(f"📋 Vérification onglet {onglet_tft}:")
        
        # Chercher une ligne avec REF (ex: ZA)
        ref_trouve = False
        for row in ws_tft.iter_rows(min_row=5, max_row=50, min_col=1, max_col=1):
            cell = row[0]
            if cell.value == 'ZA':  # Trésorerie ouverture
                row_num = cell.row
                ref_trouve = True
                
                # Lire les colonnes I et K
                val_n = ws_tft[f'I{row_num}'].value
                val_n1 = ws_tft[f'K{row_num}'].value
                
                print(f"   Ligne ZA (Trésorerie ouverture) - Ligne {row_num}:")
                print(f"      Colonne I (N): {val_n}")
                print(f"      Colonne K (N-1): {val_n1}")
                
                if val_n is not None and val_n != 0:
                    print(f"   ✅ PROBLÈME 3 RÉSOLU: TFT rempli!")
                else:
                    print(f"   ⚠️  PROBLÈME 3 PERSISTE: TFT vide")
                
                break
        
        if not ref_trouve:
            print(f"   ⚠️  REF 'ZA' non trouvée dans l'onglet TFT")
    else:
        print(f"   ⚠️  Onglet TFT non trouvé")
    
    wb.close()
    
except Exception as e:
    print(f"❌ ERREUR lors de la vérification: {e}")
    import traceback
    traceback.print_exc()

print()
print("=" * 80)
print("TEST TERMINÉ")
print("=" * 80)
print()
print(f"📄 Fichier généré: {output_file}")
print(f"   Ouvrez ce fichier pour vérifier manuellement:")
print(f"   - Onglet ACTIF: colonnes E, F, G, H")
print(f"   - Onglet TFT: colonnes I, K")
print()
